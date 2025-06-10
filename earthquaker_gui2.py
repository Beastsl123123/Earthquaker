import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
import streamlit as st
from io import BytesIO


def get_data(startdate, enddate):
    url = "https://earthquake.usgs.gov/fdsnws/event/1/query"
    params = {
        "format": "geojson",
        "starttime": startdate,
        "endtime": enddate,
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        st.error(f"Failed to fetch data, status code: {response.status_code}")
        return None


def extract_data(raw_data):
    cleaned_data = []
    features = raw_data.get("features", [])
    for feature in features:
        earthquake = {
            "Magnitude": feature["properties"]["mag"],
            "Location": feature["properties"]["place"],
            "Latitude": feature["geometry"]["coordinates"][1],
            "Longitude": feature["geometry"]["coordinates"][0],
            "Time (UTC)": datetime.utcfromtimestamp(feature["properties"]["time"] / 1000).strftime(
                "%Y-%m-%d %H:%M:%S"),
        }
        cleaned_data.append(earthquake)
    return cleaned_data


def df_to_excel(df):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter

    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Earthquakes"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="000000", size=12)
                cell.fill = PatternFill("solid", fgColor="EDEDED")
                cell.alignment = Alignment(horizontal="center")
            else:
                if c_idx == 1 and value is not None:
                    if value >= 5.0:
                        cell.fill = PatternFill("solid", fgColor="FFDBDB")
                        cell.font = Font(color="9B1C1C", bold=True)
                    elif value >= 3.0:
                        cell.fill = PatternFill("solid", fgColor="FFF4DB")
                        cell.font = Font(color="9B6A1C")
                    else:
                        cell.fill = PatternFill("solid", fgColor="F7F7F7")
                elif c_idx in [3, 4, 5]:
                    cell.alignment = Alignment(horizontal="center")

    widths = [12, 50, 12, 12, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(output)
    output.seek(0)
    return output


def main():
    st.set_page_config(
        page_title="Earthquake Data Viewer",
        page_icon="🌍",
        layout="centered",
        initial_sidebar_state="auto",
    )

    # Custom minimal styling for sleek MacOS-like UI
    st.markdown(
        """
        <style>
            /* Container max width */
            .main .block-container {
                max-width: 900px;
                padding-top: 4rem;
                padding-bottom: 4rem;
                background-color: #ffffff;
            }
            /* Headline */
            h1 {
                font-weight: 700;
                font-size: 48px;
                margin-bottom: 1rem;
                color: #000000;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen,
                    Ubuntu, Cantarell, "Open Sans", "Helvetica Neue", sans-serif;
            }
            /* Body text */
            .css-1d391kg p, .css-1d391kg label {
                color: #6b7280 !important;
                font-size: 17px;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen,
                    Ubuntu, Cantarell, "Open Sans", "Helvetica Neue", sans-serif;
            }
            /* Button style */
            div.stButton > button {
                background-color: #000000;
                color: white;
                font-weight: 700;
                padding: 0.75rem 1.75rem;
                border-radius: 12px;
                border: none;
                text-transform: uppercase;
                font-size: 14px;
                letter-spacing: 0.05em;
                transition: background-color 0.25s cubic-bezier(0.4, 0, 0.2, 1);
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
            }
            div.stButton > button:hover {
                background-color: #333333;
                cursor: pointer;
            }
            /* Cards */
            .card {
                background: #ffffff;
                border-radius: 0.75rem;
                box-shadow: 0 4px 14px rgba(0,0,0,0.07);
                padding: 1.5rem 2rem;
                margin-bottom: 1.5rem;
            }
            .card h2 {
                font-weight: 700;
                font-size: 22px;
                margin-bottom: 0.25rem;
                color: #111111;
            }
            .card p {
                font-weight: 600;
                font-size: 32px;
                color: #111111;
                margin: 0;
                line-height: 1;
            }
            /* Table styling */
            .stDataFrame table {
                border-collapse: separate;
                border-spacing: 0 12px;
                width: 100%;
                font-size: 16px;
                color: #4b5563;
                font-weight: 500;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Oxygen,
                    Ubuntu, Cantarell, "Open Sans", "Helvetica Neue", sans-serif;
            }
            .stDataFrame thead th {
                background-color: #f3f4f6;
                color: #374151;
                font-weight: 700;
                font-size: 16px;
                padding: 12px 16px;
                border-radius: 10px 10px 0 0;
                text-align: center;
            }
            .stDataFrame tbody tr:hover {
                background-color: #ede9fe;
            }
            .stDataFrame tbody td {
                background-color: #ffffff;
                padding: 14px 16px;
                border-radius: 10px;
                text-align: center;
            }
            .magnitude-high {
                background-color: #fee2e2 !important;
                color: #991b1b !important;
                font-weight: 700 !important;
            }
            .magnitude-mid {
                background-color: #fef3c7 !important;
                color: #92400e !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("<h1>Earthquake Data Viewer</h1>", unsafe_allow_html=True)
    st.markdown(
        "Analyze recent earthquake data with ease. Select your date range and review quake details below."
    )

    utc_now = datetime.now(timezone.utc)
    default_start = (utc_now - timedelta(days=1)).date()
    default_end = utc_now.date()

    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input(
            "Start Date", value=default_start, max_value=default_end
        )
    with col2:
        end_date = st.date_input("End Date", value=default_end, min_value=start_date)

    if start_date > end_date:
        st.error("Error: Start date must be before or equal to end date.")

    if st.button("Fetch Earthquake Data") and start_date <= end_date:
        start_iso = datetime.combine(start_date, datetime.min.time()).replace(
            tzinfo=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S")
        end_iso = datetime.combine(end_date, datetime.max.time()).replace(
            tzinfo=timezone.utc
        ).strftime("%Y-%m-%dT%H:%M:%S")

        with st.spinner("Fetching earthquake data..."):
            raw_data = get_data(start_iso, end_iso)

        if not raw_data or not raw_data.get("features"):
            st.warning("No earthquake data found for the selected date range.")
            return

        data = extract_data(raw_data)
        df = pd.DataFrame(data)

        total_earthquakes = len(df)
        max_magnitude = df["Magnitude"].max()
        avg_magnitude = df["Magnitude"].mean()

        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown(
                f'<div class="card"><h2>Total Earthquakes</h2><p>{total_earthquakes}</p></div>',
                unsafe_allow_html=True,
            )
        with col2:
            st.markdown(
                f'<div class="card"><h2>Max Magnitude</h2><p>{max_magnitude:.2f}</p></div>',
                unsafe_allow_html=True,
            )
        with col3:
            st.markdown(
                f'<div class="card"><h2>Avg Magnitude</h2><p>{avg_magnitude:.2f}</p></div>',
                unsafe_allow_html=True,
            )

        st.markdown("### Earthquake Details")

        def style_df(df):
            def magnitude_style(val):
                if val >= 5.0:
                    return "background-color: #fee2e2; color:#991b1b; font-weight:700;"
                elif val >= 3.0:
                    return "background-color: #fef3c7; color:#92400e;"
                return ""

            styled = df.style.applymap(magnitude_style, subset=["Magnitude"]) \
                .set_properties(**{'text-align': 'center'}, subset=['Latitude', 'Longitude', 'Time (UTC)', 'Magnitude'])
            return styled

        styled_df = style_df(df)
        st.dataframe(styled_df, use_container_width=True)

        csv_data = df.to_csv(index=False).encode("utf-8")
        excel_data = df_to_excel(df).getvalue()

        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name=f"earthquakes_{start_date}_{end_date}.csv",
            mime="text/csv",
        )
        st.download_button(
            label="Download Styled Excel",
            data=excel_data,
            file_name=f"earthquakes_{start_date}_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()



